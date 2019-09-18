import csv
import errno
import os
import shutil
import sys
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook


def main(argv):
    """
    File to create a filtered csv file with log entries from the excel sheets from the routeros_api_snooper
    """
    honeypots = ["australia", "brazil", "china-hk", "india", "netherlands", "us-central"]
    days = argv
    logs_dir = "/home/collector/logs/"
    log_captures_dir = "log_captures/"
    dest_dir = "daily_logs/"
    start_log_id = "7D" # change this number to the first log id present in the snapshot (hexadecimal)
    skippable_entries_message = ["user api logged in from 192.168.56.1 via api",
                                 "user api logged out from 192.168.56.1 via api"] # example of log entry that can be skipped
    skippable_entries_time_message = [["10:18:04", "user admin logged out from 192.168.56.1 via web"]] # example of log entry that can be skipped with timestamp

    for honeypot in honeypots:
        honeypot_upper = honeypot.upper()
        logs_dir_honeypot = logs_dir + honeypot + "/" + log_captures_dir
        for day in days:
            print("Creating daily logs for " + honeypot + " " + day)
            daily_logs = [['id', 'time', 'topics', 'message']]
            current_log_id = start_log_id
            source_filename = 'logs-' + honeypot_upper + day
            dest_file_location = logs_dir + honeypot + "/" + dest_dir + "daily_logs_" + honeypot + "_" + day + '.csv'

            # retrieve the files in the directory and filter on the day
            logs_honeypot_dir_file_list = os.listdir(path=logs_dir_honeypot)
            logs_honeypot_dir_file_list_day = [i for i in logs_honeypot_dir_file_list if i.startswith(source_filename)]
            logs_honeypot_dir_file_list_day.sort()

            if len(logs_honeypot_dir_file_list_day) == 0:
                continue

            # print(logs_honeypot_dir_file_list_day)
            for file in logs_honeypot_dir_file_list_day:
                try:
                    xlsx_file = load_workbook(logs_dir_honeypot + file)
                except Exception:
                    continue
                logs_sheet = xlsx_file['logs']
                row_count = logs_sheet.max_row
                file_time = file[len(source_filename)+1:len(source_filename)+9]
                # receive index of column names
                column_names = [logs_sheet['A1'].value, logs_sheet['B1'].value, logs_sheet['C1'].value, logs_sheet['D1'].value]
                id_column = column_names.index('.id')
                time_column = column_names.index('time')
                topics_column = column_names.index('topics')
                message_column = column_names.index('message')
                last_log_entry_time = logs_sheet[row_count][time_column].value
                last_log_id = logs_sheet[row_count][id_column].value[1:]
                # reset the current_log_id if the current_log number exceeds the number of log entries in the next file
                if int(current_log_id, 16) > int(last_log_id, 16):
                    current_log_id = start_log_id
                time_difference = convert_time(last_log_entry_time, file_time)
                for row in logs_sheet['A2':'D' + str(row_count)]:
                    # retrieve entry from the row and check if it needs to be copied
                    row_id = str(row[id_column].value)[1:]
                    # if id smaller than current log id, skip entry
                    if int(row_id, 16) < int(current_log_id, 16):
                        continue
                    # increase the current_log_id to the id of this row
                    current_log_id = row_id
                    row_message = row[message_column].value
                    row_time = row[time_column].value
                    row_topics = row[topics_column].value
                    # skip unnecessary messages
                    if row_message in skippable_entries_message or \
                            [row_time, row_message] in skippable_entries_time_message:
                        continue
                    row_time = convert_time(row_time, time_difference)
                    # add entry to the logs
                    daily_logs_entry = [row_id, row_time, row_topics, row_message]
                    daily_logs.append(daily_logs_entry)
            with open(dest_file_location, 'w') as f:
                print("writing logs to: " + dest_file_location)
                writer = csv.writer(f)
                writer.writerows(daily_logs)
            # copy files to archive folder
            print("copy files to archive folder")
            for f in logs_honeypot_dir_file_list_day:
                path = logs_dir + honeypot + "/" + log_captures_dir
                path_day = path + day + "/"
                try:
                    os.makedirs(path_day)
                except OSError as exc:  # Python >2.5
                    if exc.errno == errno.EEXIST and os.path.isdir(path_day):
                        pass
                    else:
                        raise
                shutil.move(path + f, path_day + f)


def convert_time(t1, t2) -> str:
    """
    Subtracts to timestamps and creates a string in the HH:MM:SS format with the new time
    :param t1: the time t2 will be subtracted from
    :param t2: the time to subtract from t1
    :return: a time in the HH:MM:SS format
    """
    date_time_format = "%H:%M:%S"
    other_date_time_format = "%b/%d %H:%M:%S"
    t1_time_format = date_time_format if len(t1) <= 8 else other_date_time_format
    t2_time_format = date_time_format if len(t2) <= 8 else other_date_time_format
    tdelta = datetime.strptime(t1, t1_time_format) - datetime.strptime(t2, t2_time_format)
    if tdelta.days < 0 or tdelta.days > 0:
        tdelta = timedelta(days=0, seconds=tdelta.seconds,
                           microseconds=tdelta.microseconds)
    return str(tdelta)


if __name__ == "__main__":
    main(sys.argv[1:])
